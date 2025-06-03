from ttkthemes import ThemedTk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import pandas as pd
from openpyxl.utils import get_column_letter

# --------------------------
# Función para ordenar el Treeview al hacer clic en el encabezado
# --------------------------
def treeview_sort_column(tv, col, reverse):
    # Se obtiene una lista de tuplas (valor de la columna, iid)
    l = [(tv.set(k, col), k) for k in tv.get_children('')]
    try:
        # Intenta convertir a float en caso de que sean valores numéricos
        l.sort(key=lambda t: float(t[0]) if t[0].replace('.', '', 1).isdigit() else t[0],
               reverse=reverse)
    except Exception:
        l.sort(key=lambda t: t[0], reverse=reverse)
    # Reorganiza los items en el Treeview de acuerdo al orden obtenido
    for index, (val, k) in enumerate(l):
        tv.move(k, '', index)
    # Cambia el comando del encabezado para que al siguiente clic invierta el orden
    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))


# --------------------------
# Configuración inicial y ventana con tema moderno
# --------------------------
ventana = ThemedTk(theme="arc")
ventana.title("DyadInventory")
ventana.geometry("950x600")

# Configurar estilos para ttk
style = ttk.Style(ventana)
style.configure("TButton", font=("Helvetica", 12, "bold"), padding=6)
style.map("TButton", foreground=[("active", "darkblue")], background=[("active", "#e0e0e0")])
style.configure("CRUD.TButton", font=("Helvetica", 10, "bold"), padding=6)
style.map("CRUD.TButton", foreground=[("active", "darkblue")], background=[("active", "#e0e0e0")])
style.configure("CRUD.TLabel", font=("Helvetica", 10))
style.configure("CRUD.TEntry", font=("Helvetica", 10))
style.configure("TLabel", font=("Helvetica", 12))
style.configure("TEntry", font=("Helvetica", 12))

# --------------------------
# Encabezado de la ventana
# --------------------------
header_frame = tk.Frame(ventana, bg="#2c3e50", padx=10, pady=10)
header_frame.pack(side="top", fill="x")
header_label = tk.Label(header_frame,
                        text="DyadInventory",
                        font=("Helvetica", 20, "bold"),
                        bg="#2c3e50", fg="white")
header_label.pack()

# --------------------------
# División de la ventana: sidebar y área de contenidos
# --------------------------
frame_sidebar = tk.Frame(ventana, padx=10, pady=10, bg="#ecf0f1")
frame_sidebar.pack(side="left", fill="y")
frame_tabla = tk.Frame(ventana, padx=10, pady=10)
frame_tabla.pack(side="right", fill="both", expand=True)
default_bg = frame_sidebar.cget("bg")  # Fondo por defecto

# Crear íconos para el menú contextual
icon_no_fill = tk.PhotoImage(width=16, height=16)
icon_no_fill.put(default_bg, to=(0, 0, 16, 16))
icon_green = tk.PhotoImage(width=16, height=16)
icon_green.put("lightgreen", to=(0, 0, 16, 16))
icon_purple = tk.PhotoImage(width=16, height=16)
icon_purple.put("#9370DB", to=(0, 0, 16, 16))

excel_importado = False

# --------------------------
# Variables de estado y funciones auxiliares
# --------------------------
default_text = "Ingresa un serial y presiona Enter"
default_fg = "black"

def revert_status():
    etiqueta_estado.config(text=default_text, fg=default_fg, bg=default_bg)

ruta_db = "C:/Users/Rob Vinz/Documents/Proyectos/productos.db"  # Ajusta la ruta según tu entorno
columnas = []          # Nombres de columnas de la tabla "inventario"
columnas_visibles = [] # Copia de los nombres de columnas
opciones_columnas = {} # Diccionario: {columna: BooleanVar}
resaltados_guardados = {}  # Diccionario: {rowid: tag}

# --------------------------
# Funciones para el control de checkbuttons (filtros)
# --------------------------
def marcar_todo():
    for col in opciones_columnas:
        opciones_columnas[col].set(True)

def desmarcar_todo():
    for col in opciones_columnas:
        opciones_columnas[col].set(False)

def aplicar_filtros():
    actualizar_treeview_columnas()

# --------------------------
# Widgets del Sidebar Superior
# --------------------------
entrada_serial = ttk.Entry(frame_sidebar)
entrada_serial.pack(pady=10, fill="x")
etiqueta_estado = tk.Label(frame_sidebar, text=default_text, fg=default_fg, bg=default_bg, font=("Helvetica", 10))
etiqueta_estado.pack(pady=(0,5), fill="x")
boton_importar = ttk.Button(frame_sidebar, text="Importar Excel", style="CRUD.TButton")
boton_importar.pack(pady=5, fill="x")
boton_exportar = ttk.Button(frame_sidebar, text="Exportar Excel", style="CRUD.TButton")
boton_exportar.pack(pady=5, fill="x")

# --------------------------
# Sección de Operaciones CRUD en el Sidebar
# --------------------------
frame_crud = tk.LabelFrame(frame_sidebar, text="Operaciones CRUD", bg=default_bg, font=("Helvetica", 10))
frame_crud.pack(pady=10, fill="x")
btn_agregar = ttk.Button(frame_crud, text="Agregar registro", style="CRUD.TButton",
                         command=lambda: add_record(), state="disabled")
btn_agregar.pack(pady=2, fill="x")
btn_editar = ttk.Button(frame_crud, text="Editar registro", style="CRUD.TButton",
                        command=lambda: edit_record(), state="disabled")
btn_editar.pack(pady=2, fill="x")
btn_eliminar = ttk.Button(frame_crud, text="Eliminar registro", style="CRUD.TButton",
                          command=lambda: delete_record(), state="disabled")
btn_eliminar.pack(pady=2, fill="x")

# --------------------------
# Contenedor para Filtros (checkbuttons)
# --------------------------
checkbuttons_container = tk.Frame(frame_sidebar, bg="#ecf0f1")
checkbuttons_container.pack(pady=(10,0), anchor="w", fill="x")
canvas_checkbuttons = tk.Canvas(checkbuttons_container, bg="#ecf0f1", highlightthickness=0)
canvas_checkbuttons.pack(side="left", fill="both", expand=True)
scrollbar_checkbuttons = tk.Scrollbar(checkbuttons_container, orient="vertical", command=canvas_checkbuttons.yview)
scrollbar_checkbuttons.pack(side="right", fill="y")
canvas_checkbuttons.configure(yscrollcommand=scrollbar_checkbuttons.set)
checkbuttons_frame = tk.Frame(canvas_checkbuttons, bg="#ecf0f1")
checkbuttons_frame_window = canvas_checkbuttons.create_window((0, 0), window=checkbuttons_frame, anchor="nw")

def on_checkbuttons_configure(event):
    canvas_checkbuttons.configure(scrollregion=canvas_checkbuttons.bbox("all"))
    canvas_checkbuttons.itemconfig(checkbuttons_frame_window, width=canvas_checkbuttons.winfo_width())
checkbuttons_frame.bind("<Configure>", on_checkbuttons_configure)

def _on_mousewheel(event):
    canvas_checkbuttons.yview_scroll(-3 * int(event.delta/120), "units")
for widget in (canvas_checkbuttons, checkbuttons_container, checkbuttons_frame):
    widget.bind("<MouseWheel>", _on_mousewheel)

def on_sidebar_configure(event):
    new_height = max(300, int(event.height * 0.6))
    canvas_checkbuttons.config(height=new_height)
frame_sidebar.bind("<Configure>", on_sidebar_configure)

def reconstruir_checkbuttons():
    global opciones_columnas, control_buttons, excel_importado
    for widget in checkbuttons_frame.winfo_children():
        widget.destroy()
    opciones_columnas.clear()
    for col in columnas:
        opciones_columnas[col] = tk.BooleanVar(value=True)
        chk = tk.Checkbutton(checkbuttons_frame, text=col, variable=opciones_columnas[col],
                             bg="#ecf0f1", font=("Helvetica", 10))
        chk.pack(anchor="w")
    if excel_importado:
        btn_marcar = ttk.Button(checkbuttons_frame, text="Marcar todas las casillas", 
                                command=marcar_todo, style="CRUD.TButton")
        btn_marcar.pack(pady=(5,0), fill="x", padx=10)
        btn_desmarcar = ttk.Button(checkbuttons_frame, text="Desmarcar todas las casillas", 
                                   command=desmarcar_todo, style="CRUD.TButton")
        btn_desmarcar.pack(pady=(0,5), fill="x", padx=10)
        btn_aplicar = ttk.Button(checkbuttons_frame, text="Aplicar", 
                                 command=aplicar_filtros, style="CRUD.TButton")
        btn_aplicar.pack(pady=(0,5), fill="x", padx=10)
        control_buttons = [btn_marcar, btn_desmarcar, btn_aplicar]
    checkbuttons_frame.update_idletasks()
    canvas_checkbuttons.configure(scrollregion=canvas_checkbuttons.bbox("all"))

# --------------------------
# Treeview y Scrollbars en el Área de Contenidos
# --------------------------
scroll_x = tk.Scrollbar(frame_tabla, orient="horizontal", width=25, relief="solid")
scroll_y = tk.Scrollbar(frame_tabla, orient="vertical", width=25, relief="solid")
tabla = ttk.Treeview(frame_tabla, show="headings")
scroll_x.config(command=tabla.xview)
scroll_y.config(command=tabla.yview)
scroll_x.pack(side="bottom", fill="x")
scroll_y.pack(side="right", fill="y")
tabla.tag_configure("resaltado", background="lightgreen")
tabla.tag_configure("duplicado", background="#9370DB", foreground="white")
tabla.tag_configure("morado", background="#9370DB", foreground="white")
tabla.tag_configure("cambiado", background="yellow")

# Función para un scroll más fluido en el Treeview (con saltos reducidos)
def _on_treeview_mousewheel(event):
    current_view = tabla.yview()[0]
    step = 0.01 * (event.delta / 120)
    new_view = current_view - step
    new_view = max(0, min(new_view, 1))
    tabla.yview_moveto(new_view)

tabla.bind("<MouseWheel>", _on_treeview_mousewheel)

# --------------------------
# Menú contextual para el Treeview con iconos de colores
# --------------------------
def quitar_color(row_id):
    tabla.item(row_id, tags=())
    if row_id in resaltados_guardados:
        del resaltados_guardados[row_id]
    print(f"Fila {row_id} sin relleno")

def poner_verde(row_id):
    tabla.item(row_id, tags=("resaltado",))
    resaltados_guardados[row_id] = "resaltado"
    print(f"Fila {row_id} en verde")

def poner_morado(row_id):
    tabla.item(row_id, tags=("morado",))
    resaltados_guardados[row_id] = "morado"
    print(f"Fila {row_id} en morado")

def on_right_click(event):
    print(f"Click derecho detectado en x={event.x}, y={event.y}")
    row_id = tabla.identify_row(event.y)
    print("Fila identificada:", row_id)
    if row_id:
        tabla.selection_set(row_id)
        context_menu.entryconfigure("Sin relleno", command=lambda: quitar_color(row_id))
        context_menu.entryconfigure("Verde", command=lambda: poner_verde(row_id))
        context_menu.entryconfigure("Morado", command=lambda: poner_morado(row_id))
        context_menu.post(event.x_root, event.y_root)
    else:
        print("No se identificó ninguna fila en esa posición.")

context_menu = tk.Menu(ventana, tearoff=0)
context_menu.add_command(label="Sin relleno", image=icon_no_fill, compound="left")
context_menu.add_command(label="Verde", image=icon_green, compound="left")
context_menu.add_command(label="Morado", image=icon_purple, compound="left")

# --------------------------
# Funciones CRUD con ventanas responsive y scrollable para Agregar y Editar
# --------------------------
def add_record():
    top = tk.Toplevel(ventana)
    top.title("Agregar registro")
    canvas = tk.Canvas(top)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar = tk.Scrollbar(top, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)
    top.bind("<MouseWheel>", lambda event: canvas.yview_scroll(-1 * int(event.delta/120), "units"))
    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    
    entries = {}
    for i, col in enumerate(columnas):
        lbl = ttk.Label(scrollable_frame, text=col, style="CRUD.TLabel", width=15, anchor="w")
        lbl.grid(row=i, column=0, padx=5, pady=5, sticky="w")
        ent = ttk.Entry(scrollable_frame, style="CRUD.TEntry")
        ent.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
        entries[col] = ent
    scrollable_frame.grid_columnconfigure(1, weight=1)
    top.update_idletasks()
    req_width = top.winfo_reqwidth()
    req_height = max(top.winfo_reqheight(), 600)
    top.geometry(f"{req_width}x{req_height}")
    
    def save_new():
        vals = [entries[col].get() for col in columnas]
        conn = sqlite3.connect(ruta_db)
        cursor = conn.cursor()
        query = "INSERT INTO inventario (" + ",".join(f'"{col}"' for col in columnas) + ") VALUES (" + ",".join("?" for _ in columnas) + ")"
        cursor.execute(query, vals)
        conn.commit()
        conn.close()
        top.destroy()
        actualizar_estructura()
        actualizar_treeview_columnas()
    btn_save = ttk.Button(scrollable_frame, text="Guardar", style="CRUD.TButton", command=save_new)
    btn_save.grid(row=len(columnas), column=0, columnspan=2, pady=10)

def edit_record():
    selected = tabla.focus()
    if not selected:
        messagebox.showwarning("Editar registro", "Por favor, seleccione un registro para editar.")
        return
    current_values = tabla.item(selected, "values")
    top = tk.Toplevel(ventana)
    top.title("Editar registro")
    
    canvas = tk.Canvas(top)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar = tk.Scrollbar(top, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)
    top.bind("<MouseWheel>", lambda event: canvas.yview_scroll(-1 * int(event.delta/120), "units"))
    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    
    entries = {}
    for i, col in enumerate(columnas):
        lbl = ttk.Label(scrollable_frame, text=col, style="CRUD.TLabel", width=15, anchor="w")
        lbl.grid(row=i, column=0, padx=5, pady=5, sticky="w")
        ent = ttk.Entry(scrollable_frame, style="CRUD.TEntry")
        ent.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
        if i < len(current_values):
            ent.insert(0, current_values[i])
        entries[col] = ent
    scrollable_frame.grid_columnconfigure(1, weight=1)
    top.update_idletasks()
    req_width = top.winfo_reqwidth()
    req_height = max(top.winfo_reqheight(), 600)
    top.geometry(f"{req_width}x{req_height}")
    
    def save_edit():
        new_vals = [entries[col].get() for col in columnas]
        changed = any(orig != new for orig, new in zip(current_values, new_vals))
        conn = sqlite3.connect(ruta_db)
        cursor = conn.cursor()
        sets = ",".join(f'"{col}"=?' for col in columnas)
        query = f"UPDATE inventario SET {sets} WHERE rowid = ?"
        cursor.execute(query, new_vals + [selected])
        conn.commit()
        conn.close()
        top.destroy()
        if changed:
            resaltados_guardados[selected] = "cambiado"
        actualizar_treeview_columnas()
        if selected in tabla.get_children():
            tabla.selection_set(selected)
            tabla.focus(selected)
    btn_save = ttk.Button(scrollable_frame, text="Guardar cambios", style="CRUD.TButton", command=save_edit)
    btn_save.grid(row=len(columnas), column=0, columnspan=2, pady=10)

def delete_record():
    selected = tabla.focus()
    if not selected:
        messagebox.showwarning("Eliminar registro", "Por favor, seleccione un registro para eliminar.")
        return
    if messagebox.askyesno("Eliminar registro", "¿Está seguro de eliminar el registro seleccionado?"):
        conn = sqlite3.connect(ruta_db)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM inventario WHERE rowid = ?", (selected,))
        conn.commit()
        conn.close()
        actualizar_estructura()
        actualizar_treeview_columnas()

btn_agregar.config(command=add_record)
btn_editar.config(command=edit_record)
btn_eliminar.config(command=delete_record)

# --------------------------
# Funciones para estructurar y manejar datos (Treeview y filtros)
# --------------------------
def actualizar_treeview_columnas():
    global tabla
    tabla.destroy()
    columnas_seleccionadas = [col for col in columnas if opciones_columnas.get(col, tk.BooleanVar(value=True)).get()]
    if not columnas_seleccionadas:
        columnas_seleccionadas = columnas.copy()
    tabla_nueva = ttk.Treeview(frame_tabla, show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
    tabla_nueva["columns"] = columnas_seleccionadas
    for col in columnas_seleccionadas:
        # Aquí se añade el comando en el encabezado para ordenar la columna
        tabla_nueva.heading(col, text=col, anchor="w", 
                              command=lambda c=col: treeview_sort_column(tabla_nueva, c, False))
        tabla_nueva.column(col, width=150, anchor="w")
    tabla_nueva.pack(expand=True, fill="both")
    tabla = tabla_nueva
    scroll_x.config(command=tabla.xview)
    scroll_y.config(command=tabla.yview)
    tabla.tag_configure("resaltado", background="lightgreen")
    tabla.tag_configure("duplicado", background="#9370DB", foreground="white")
    tabla.tag_configure("morado", background="#9370DB", foreground="white")
    tabla.tag_configure("cambiado", background="yellow")
    tabla.bind("<MouseWheel>", _on_treeview_mousewheel)
    cargar_inventario()

def cargar_inventario():
    conn = sqlite3.connect(ruta_db)
    cursor = conn.cursor()
    cursor.execute("SELECT rowid, * FROM inventario")
    registros = cursor.fetchall()
    conn.close()
    tabla.delete(*tabla.get_children())
    columnas_seleccionadas = ([col for col in columnas if opciones_columnas.get(col, tk.BooleanVar(value=True)).get()]
                              if opciones_columnas else columnas.copy())
    for index, fila in enumerate(registros):
        uid = str(fila[0])
        datos_filtrados = [fila[columnas.index(col)+1] for col in columnas_seleccionadas]
        tag = "even" if index % 2 == 0 else "odd"
        if uid in resaltados_guardados:
            tag = resaltados_guardados[uid]
        tabla.insert("", "end", iid=uid, values=tuple(datos_filtrados), tags=(tag,))
    
    tabla.tag_configure("even", background="white")
    tabla.tag_configure("odd", background="#f0f0f0")
    tabla.tag_configure("cambiado", background="yellow")

def actualizar_estructura():
    global columnas, columnas_visibles
    conn = sqlite3.connect(ruta_db)
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(inventario)")
    columnas = [col[1] for col in cursor.fetchall()]
    conn.close()
    columnas_visibles = columnas.copy()
    print("Columnas actualizadas:", columnas)

def buscar_serial():
    serial_buscado = entrada_serial.get().strip().upper()
    encontrado = False
    first_found = None
    for item in tabla.get_children():
        valores = tuple(str(v).upper() for v in tabla.item(item, "values"))
        if serial_buscado in valores:
            if first_found is None:
                first_found = item
            estado_actual = resaltados_guardados.get(item, None)
            if estado_actual is None:
                tabla.item(item, tags=("resaltado",))
                resaltados_guardados[item] = "resaltado"
            elif estado_actual == "resaltado":
                tabla.item(item, tags=("duplicado",))
                resaltados_guardados[item] = "duplicado"
            encontrado = True
    if first_found:
        tabla.see(first_found)
    if not encontrado:
        etiqueta_estado.config(text="❌ Serial no encontrado.", fg="red", bg=default_bg)
    else:
        etiqueta_estado.config(text="✅ Serial(es) resaltado(s)/duplicado(s).", fg="green", bg=default_bg)
        entrada_serial.delete(0, tk.END)
    ventana.after(5000, revert_status)

def importar_excel():
    global excel_importado
    file_path = filedialog.askopenfilename(title="Selecciona un archivo de Excel",
                                           filetypes=[("Archivos de Excel", "*.xls;*.xlsx")])
    if file_path:
        try:
            df = pd.read_excel(file_path)
            conn = sqlite3.connect(ruta_db)
            df.to_sql("inventario", conn, if_exists="replace", index=False)
            conn.close()
            resaltados_guardados.clear()
            excel_importado = True
            btn_agregar.config(state='normal')
            btn_editar.config(state='normal')
            btn_eliminar.config(state='normal')
            etiqueta_estado.config(text="Inventario importado exitosamente!", fg="green", bg=default_bg)
            actualizar_estructura()
            reconstruir_checkbuttons()
            actualizar_treeview_columnas()
            ventana.after(5000, revert_status)
        except Exception as e:
            etiqueta_estado.config(text=f"Error al importar Excel: {e}", fg="red", bg=default_bg)
            ventana.after(5000, revert_status)

def exportar_excel():
    file_path = filedialog.asksaveasfilename(title="Guardar archivo Excel",
                                             defaultextension=".xlsx",
                                             filetypes=[("Archivos de Excel", "*.xlsx")])
    if not file_path:
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        columnas_export = tabla["columns"]
        for col_index, col_name in enumerate(columnas_export, start=1):
            cell = ws.cell(row=1, column=col_index, value=col_name)
            cell.font = Font(bold=True)
            tree_width = tabla.column(col_name)["width"]
            excel_width = tree_width / 7  # Aproximación: ~7 px = 1 unidad en Excel
            col_letter = get_column_letter(col_index)
            ws.column_dimensions[col_letter].width = excel_width
        row_index = 2
        for item in tabla.get_children():
            valores = tabla.item(item, "values")
            tags = tabla.item(item, "tags")
            if "resaltado" in tags:
                fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif "duplicado" in tags or "morado" in tags:
                fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")
            else:
                fill = None
            for col_index, value in enumerate(valores, start=1):
                try:
                    numeric_value = float(value)
                    if numeric_value.is_integer():
                        cell_value = int(numeric_value)
                    else:
                        cell_value = numeric_value
                except Exception:
                    cell_value = value
                cell = ws.cell(row=row_index, column=col_index, value=cell_value)
                if fill:
                    cell.fill = fill
            row_index += 1
        wb.save(file_path)
        etiqueta_estado.config(text="El archivo Excel se exportó exitosamente.", fg="green", bg=default_bg)
        ventana.after(5000, revert_status)
    except Exception as e:
        etiqueta_estado.config(text=f"Error al exportar Excel: {e}", fg="red", bg=default_bg)
        ventana.after(5000, revert_status)

# --------------------------
# Bindings y asignación de comandos
# --------------------------
ventana.bind_class("Treeview", "<Button-3>", on_right_click)
tabla.bind("<Button-3>", on_right_click)
boton_importar.config(command=importar_excel)
boton_exportar.config(command=exportar_excel)
entrada_serial.bind("<Return>", lambda event: buscar_serial())

# --------------------------
# Inicialización de la interfaz
# --------------------------
reconstruir_checkbuttons()
actualizar_estructura()
cargar_inventario()

ventana.mainloop()